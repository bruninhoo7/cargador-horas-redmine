"""
Lógica principal de carga de horas: lee Excel, procesa escenarios,
llama a Redmine y actualiza el Excel con los resultados.
"""
import pandas as pd
import openpyxl
from .redmine_api import (
    obtener_id_actividad, obtener_mi_id, crear_issue, cargar_entrada,
    obtener_titulo_issue, verificar_duplicado_redmine,
    limpiar, armar_titulo_issue, armar_comentario,
)


# Clientes especiales con issue fijo (sin importar mayúsculas)
CLIENTES_ISSUE_FIJO = {
    "licencia":   7038,
    "vacaciones": 7038,
    "vacacion":   7038,
    "vacas":      7038,
}


def ejecutar_carga(cfg, clientes_lista, log, on_done,
                   fecha_desde=None, fecha_hasta=None):
    """Función principal de carga.

    cfg            : dict con configuración (redmine_url, api_key, archivo_excel, hoja, etc.)
    clientes_lista : lista de clientes [{nombres_excel, proyecto_id}]
    log            : callback para escribir mensajes
    on_done        : callback final (ok, horas_cargadas, issues_creados)
    fecha_desde    : datetime o None — filtro inferior
    fecha_hasta    : datetime o None — filtro superior
    """
    def L(m): log(m)
    try:
        # Mapeo alias → proyecto_id (case-insensitive, soporta múltiples nombres)
        mapeo = {}
        for c in clientes_lista:
            nombres = c.get("nombres_excel", c.get("nombre_excel", ""))
            for nombre in [n.strip() for n in nombres.split(",") if n.strip()]:
                mapeo[nombre.lower()] = c["proyecto_id"]

        try:
            df = pd.read_excel(cfg["archivo_excel"], sheet_name=cfg["hoja"])
        except Exception as e:
            L(f"❌ No se pudo leer el Excel: {e}")
            on_done(False)
            return

        total_filas = len(df)

        if fecha_desde and fecha_hasta:
            df["_fecha_dt"] = pd.to_datetime(df["Fecha"], errors="coerce")
            df = df[(df["_fecha_dt"] >= fecha_desde) & (df["_fecha_dt"] <= fecha_hasta)]
            df = df.drop(columns=["_fecha_dt"])
            L(f"Procesando {len(df)} fila(s) en el rango seleccionado...")
        else:
            L(f"Procesando {total_filas} fila(s)...")

        act_id, err = obtener_id_actividad(
            cfg["redmine_url"], cfg["api_key"], cfg["actividad"])
        if err:
            L(f"❌ Error de configuración: {err}")
            on_done(False); return

        usr_id, err = obtener_mi_id(cfg["redmine_url"], cfg["api_key"])
        if err:
            L(f"❌ Error de autenticación: {err}")
            on_done(False); return

        ok = errores = omitidas = duplicados = creados = 0
        cache = {}                       # (proyecto_id, titulo_issue) → id_redmine
        ids_actualizados = {}            # idx → nuevo id_redmine
        id_tickets_actualizados = {}     # idx → (id_ticket, titulo) de Redmine
        comentarios_actualizados = {}    # idx → comentario auto-generado

        usa_id_ticket = cfg.get("usa_id_ticket", True)

        for idx, row in df.iterrows():
            fn               = idx + 2
            id_ticket        = row.get("ID_Ticket") if usa_id_ticket else None
            titulo_col       = row.get("Titulo")
            comentario_col   = row.get("Comentario")
            comentario_carga = armar_comentario(comentario_col, id_ticket, titulo_col)

            if not limpiar(comentario_col) and comentario_carga:
                comentarios_actualizados[idx] = comentario_carga

            id_redmine = row.get("ID_Redmine")
            tiene_id   = not (pd.isna(id_redmine) or str(id_redmine).strip() == "")

            # --- ESCENARIO A: Tiene ID → cargar horas ---
            if tiene_id:
                id_ticket_actual = limpiar(id_ticket)
                titulo_actual    = limpiar(titulo_col)
                if not id_ticket_actual or not titulo_actual:
                    id_ticket_rd, titulo_rd = obtener_titulo_issue(
                        cfg["redmine_url"], cfg["api_key"], id_redmine)
                    if id_ticket_rd is not None or titulo_rd is not None:
                        val_ticket = id_ticket_rd if not id_ticket_actual else id_ticket_actual
                        val_titulo = titulo_rd    if not titulo_actual    else titulo_actual
                        id_tickets_actualizados[idx] = (val_ticket, val_titulo)

            # --- ESCENARIO B: Sin ID → crear issue o usar issue fijo ---
            else:
                cliente = str(row.get("Cliente", "")).strip()
                cliente_lower = cliente.lower()

                if cliente_lower in CLIENTES_ISSUE_FIJO:
                    id_redmine = CLIENTES_ISSUE_FIJO[cliente_lower]
                    L(f"  Fila {fn}: Cliente especial '{cliente}' → Issue fijo #{id_redmine}")
                else:
                    proyecto_id = mapeo.get(cliente_lower)
                    if proyecto_id is None:
                        ident = limpiar(titulo_col) or limpiar(id_ticket) or str(fn)
                        L(f"❌  Fila {fn} — {ident[:35]}: "
                          f"Cliente '{cliente}' no configurado en la pestaña Clientes")
                        omitidas += 1; continue
                    titulo_issue = armar_titulo_issue(id_ticket, titulo_col)
                    ck = (proyecto_id, titulo_issue)
                    if ck in cache:
                        id_redmine = cache[ck]
                        ids_actualizados[idx] = id_redmine
                    else:
                        nid, err = crear_issue(cfg["redmine_url"], cfg["api_key"],
                                               proyecto_id, titulo_issue, usr_id)
                        if err:
                            ident = limpiar(titulo_col) or limpiar(id_ticket) or str(fn)
                            L(f"❌  Fila {fn} — {ident[:35]}: "
                              f"Error creando issue — {err[:70]}")
                            errores += 1; continue
                        cache[ck] = nid
                        id_redmine = nid
                        creados += 1
                        ids_actualizados[idx] = nid

            # --- Procesar fecha ---
            fecha_raw = row.get("Fecha")
            try:
                fdt = pd.to_datetime(fecha_raw)
                fecha_str = fdt.strftime("%Y-%m-%d")
            except Exception:
                L(f"  Fila {fn}: ❌ Fecha inválida")
                errores += 1; continue

            horas = row.get("HS Trabajadas")
            if pd.isna(horas): horas = 0

            hs_cliente = row.get("HS Cliente", 0)
            if pd.isna(hs_cliente): hs_cliente = 0

            ubicacion = (cfg["ubicacion_remoto"]
                         if (cfg["dia_remoto"] >= 0 and fdt.weekday() == cfg["dia_remoto"])
                         else cfg["ubicacion"])

            # --- Verificar duplicado ---
            if verificar_duplicado_redmine(
                    cfg["redmine_url"], cfg["api_key"],
                    id_redmine, fecha_str, horas, comentario_carga):
                duplicados += 1; continue

            # --- Cargar horas ---
            status, resp = cargar_entrada(
                cfg["redmine_url"], cfg["api_key"], id_redmine, fecha_str,
                horas, comentario_carga, act_id, hs_cliente,
                cfg["tipo_hora"], ubicacion)

            if status == 201:
                ok += 1
            else:
                titulo_fila = limpiar(titulo_col) or limpiar(id_ticket) or fecha_str
                if "status" in resp.lower() and (
                        "cerrad" in resp.lower() or
                        "closed" in resp.lower() or
                        "invalid" in resp.lower()):
                    msg = ("Issue cerrado — no se pueden cargar horas. "
                           "Reabrir el issue en Redmine primero.")
                elif "422" in str(status):
                    msg = ("Datos invalidos (422) — verificar que el issue "
                           "exista y este abierto.")
                else:
                    msg = f"HTTP {status} — {resp[:60]}"
                L(f"❌  Fila {fn} — {fecha_str} — {titulo_fila[:35]}: {msg}")
                errores += 1

        # --- Actualizar Excel ---
        if ids_actualizados or id_tickets_actualizados or comentarios_actualizados:
            total_cambios = (len(ids_actualizados) + len(id_tickets_actualizados)
                             + len(comentarios_actualizados))
            L(f"\n📝 Actualizando Excel ({total_cambios} cambio(s))...")
            try:
                wb = openpyxl.load_workbook(cfg["archivo_excel"])
                ws = wb[cfg["hoja"]]
                header_row = [cell.value for cell in ws[1]]

                def col_num(nombre):
                    try: return header_row.index(nombre) + 1
                    except ValueError: return None

                col_id        = col_num("ID_Redmine")
                col_id_ticket = col_num("ID_Ticket")
                col_titulo    = col_num("Titulo")
                col_coment    = col_num("Comentario")

                if col_id:
                    for df_idx, nuevo_id in ids_actualizados.items():
                        ws.cell(row=df_idx + 2, column=col_id, value=nuevo_id)

                for df_idx, (id_ticket_rd, titulo_rd) in id_tickets_actualizados.items():
                    if col_id_ticket and id_ticket_rd:
                        ws.cell(row=df_idx + 2, column=col_id_ticket, value=id_ticket_rd)
                    if col_titulo and titulo_rd:
                        ws.cell(row=df_idx + 2, column=col_titulo, value=titulo_rd)

                for df_idx2, coment_gen in comentarios_actualizados.items():
                    if col_coment:
                        ws.cell(row=df_idx2 + 2, column=col_coment, value=coment_gen)

                wb.save(cfg["archivo_excel"])
            except Exception as e:
                L(f"❌  Error al actualizar el Excel: {e}")

        L("\n" + "=" * 50)
        L(f"  ✨ Issues creados   : {creados}")
        L(f"  ✔ Horas cargadas   : {ok}")
        L(f"  ⏭ Duplicados omit. : {duplicados}")
        L(f"  ❌ Con errores      : {errores}")
        L(f"  ⏭ Omitidas         : {omitidas}")
        L("=" * 50)
        on_done(True, ok, creados)

    except Exception as e:
        L(f"\n❌ Error inesperado: {e}")
        on_done(False, 0, 0)
