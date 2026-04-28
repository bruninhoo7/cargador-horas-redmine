"""
Cargador de Horas → Redmine  |  HM Consulting
===============================================
v1.0.0  —  Autor: HM Consulting

Requisitos:
    pip install pandas openpyxl requests pillow
"""

import os, sys, threading, json
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

# Módulos del proyecto
from core.config_storage import (
    get_app_dir, get_data_dir, get_asset,
    cargar_auth, guardar_auth, hash_pass,
    cargar_config, guardar_config,
    cargar_clientes, guardar_clientes,
    cargar_stats, guardar_stats,
    encriptar, desencriptar,
    CONFIG_DEFAULT, CRYPTO_OK,
    ACTIVIDADES, UBICACIONES, DIAS_SEMANA, TIPOS_HORA,
    CONFIG_FILE, PASS_FILE, CLIENTES_FILE, STATS_FILE,
)
from core.redmine_api import (
    obtener_proyectos, obtener_id_actividad, obtener_mi_id,
    crear_issue, cargar_entrada, obtener_titulo_issue,
    verificar_duplicado_redmine,
    limpiar, armar_titulo_issue, armar_comentario,
)
from core.updater import verificar_actualizacion, descargar_actualizacion
from core.ejecutor import ejecutar_carga
from datetime import datetime, timedelta

def verificar_instancia_unica():
    """Retorna el socket si es la primera instancia, None si ya hay una corriendo."""
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 0)
        s.bind(("127.0.0.1", 47291))
        s.listen(1)
        return s
    except OSError:
        return None

# ============================================================
#  METADATA
# ============================================================

APP_NAME    = "Cargador de Horas Redmine"
APP_VERSION = "1.5.3"
APP_AUTHOR  = "HM Consulting"

# ============================================================
#  RUTAS
# ============================================================


# ============================================================
#  COLORES
# ============================================================

AZUL_OSCURO = "#1a2a4a"
AZUL_CLARO  = "#00aadd"
BG          = "#f0f4f8"


# ============================================================
#  PANTALLA DE LOGIN
# ============================================================

class LoginScreen(tk.Tk):
    def __init__(self, auth):
        super().__init__()
        self.auth = auth
        self.title(APP_NAME)
        self.resizable(False, False)
        self.configure(bg=BG)
        ico = get_asset("HM_Icono.ico")
        if os.path.exists(ico):
            try: self.iconbitmap(ico)
            except: pass
        self._build()
        self._center()
        # Chequeo automático con delay y reintento (intentos crecientes)
        def _auto_check(intento=1):
            def worker():
                try:
                    nueva_ver, url_exe = verificar_actualizacion(APP_VERSION)
                    if nueva_ver:
                        self.after(0, lambda: self._mostrar_update(nueva_ver, url_exe))
                    elif intento < 3:
                        # Reintento en 30s y luego en 60s
                        delay = 30000 if intento == 1 else 60000
                        self.after(delay, lambda: _auto_check(intento + 1))
                except Exception:
                    if intento < 3:
                        delay = 30000 if intento == 1 else 60000
                        self.after(delay, lambda: _auto_check(intento + 1))
            threading.Thread(target=worker, daemon=True).start()
        self.after(3000, _auto_check)

    def _center(self):
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    def _build(self):
        hf = tk.Frame(self, bg=AZUL_OSCURO)
        hf.pack(fill="x")
        hf_left = tk.Frame(hf, bg="white", width=280, height=56)
        hf_left.pack(side="left")
        hf_left.pack_propagate(False)
        logo_path = get_asset("logo_app.png")
        if os.path.exists(logo_path):
            try:
                from PIL import Image, ImageTk
                img = Image.open(logo_path).convert("RGB")
                img.thumbnail((260, 48), Image.LANCZOS)
                photo = ImageTk.PhotoImage(img)
                tk.Label(hf_left, image=photo, bg="white").place(relx=0.5, rely=0.5, anchor="center")
                hf_left._photo = photo
            except: pass
        tk.Label(hf, text=f"v{APP_VERSION}", font=("Segoe UI", 9),
                 bg=AZUL_OSCURO, fg="#aaccee").pack(side="right", padx=16, pady=18)

        body = tk.Frame(self, bg=BG, padx=40, pady=20)
        body.pack(fill="both", expand=True)
        tk.Label(body, text="Ingresá tu contraseña para continuar",
                 font=("Segoe UI", 10), bg=BG, fg="#555").pack(pady=(8, 14))

        self.v_pass = tk.StringVar()
        entry = ttk.Entry(body, textvariable=self.v_pass, show="•", width=28)
        entry.pack()
        entry.focus()
        entry.bind("<Return>", lambda e: self._ingresar())

        self.lbl_error = tk.Label(body, text="", fg="#e74c3c", bg=BG, font=("Segoe UI", 8))
        self.lbl_error.pack(pady=(4, 0))

        ttk.Button(body, text="▶  Ingresar", command=self._ingresar).pack(pady=(12, 4))
        ttk.Button(body, text="¿Olvidaste tu contraseña?",
                   command=self._recuperar).pack()

    def _ingresar(self):
        if hash_pass(self.v_pass.get()) == self.auth["hash"]:
            self.destroy()
            App().mainloop()
        else:
            self.lbl_error.config(text="Contraseña incorrecta. Intentá de nuevo.")
            self.v_pass.set("")

    def _recuperar(self):
        top = tk.Toplevel(self)
        top.title("Recuperar acceso")
        top.resizable(False, False)
        top.grab_set()
        top.configure(bg=BG)
        top.geometry(f"320x200+{self.winfo_x()+50}+{self.winfo_y()+80}")
        tk.Label(top, text="Recuperar acceso",
                 font=("Segoe UI", 11, "bold"), bg=BG, fg=AZUL_OSCURO).pack(pady=(16, 4))
        tk.Label(top, text="Ingresá tu API Key para restablecer la contraseña:",
                 font=("Segoe UI", 9), bg=BG, fg="#555", justify="center").pack(pady=(4, 8))
        v_key = tk.StringVar()
        ttk.Entry(top, textvariable=v_key, width=32).pack(padx=20)
        lbl_err = tk.Label(top, text="", fg="#e74c3c", bg=BG, font=("Segoe UI", 8))
        lbl_err.pack(pady=4)

        def verificar():
            cfg = cargar_config()
            if v_key.get().strip() == cfg.get("api_key", ""):
                guardar_auth({"hash": None})
                messagebox.showinfo("Acceso restaurado",
                    "Contrasena eliminada. Podes ingresar sin contrasena y configurar una nueva en Configuracion.")
                top.destroy()
                self.destroy()
                App().mainloop()
            else:
                lbl_err.config(text="API Key incorrecta.")

        ttk.Button(top, text="Verificar", command=verificar).pack(pady=4)



# ============================================================
#  APP
# ============================================================

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"{APP_NAME}  v{APP_VERSION}")
        self.resizable(False, False)
        self.resizable(True, True)
        self.minsize(720, 560)

        ico = get_asset("HM_Icono.ico")
        if os.path.exists(ico):
            try: self.iconbitmap(ico)
            except Exception: pass

        self.cfg      = cargar_config()
        self.clientes = cargar_clientes()
        self.stats    = cargar_stats()
        self._cli_rows = []

        self._build()
        self._center()
        # Chequeo automático con delay y reintento

    def _center(self):
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    def _abrir_log(self):
        """Abre ventana de log en tiempo real."""
        if hasattr(self, "_log_win") and self._log_win.winfo_exists():
            self._log_win.lift()
            return
        win = tk.Toplevel(self)
        self._log_win = win
        win.title("Log de actividad")
        win.geometry("640x400")
        win.configure(bg=BG)
        ico = get_asset("HM_Icono.ico")
        if os.path.exists(ico):
            try: win.iconbitmap(ico)
            except: pass
        tk.Label(win, text="Log de actividad",
                 font=("Segoe UI", 10, "bold"), bg=BG, fg=AZUL_OSCURO).pack(anchor="w", padx=12, pady=(8,4))
        self._debug_box = scrolledtext.ScrolledText(
            win, font=("Consolas", 8), bg="#1e1e1e", fg="#d4d4d4",
            state="normal", width=80, height=20)
        self._debug_box.pack(fill="both", expand=True, padx=8, pady=(0,8))
        # Copiar todo el log existente
        existing = self.log_box.get("1.0", "end")
        self._debug_box.insert("end", existing)
        self._debug_box.see("end")
        bf = tk.Frame(win, bg=BG)
        bf.pack(fill="x", padx=8, pady=(0,8))
        ttk.Button(bf, text="🗑  Limpiar",
                   command=lambda: self._debug_box.delete("1.0","end")).pack(side="left")
        ttk.Button(bf, text="📋  Copiar todo",
                   command=lambda: (win.clipboard_clear(),
                                    win.clipboard_append(self._debug_box.get("1.0","end")))
                   ).pack(side="left", padx=8)

    def _check_update(self, manual=False):
        try:
            nueva_ver, url_exe = verificar_actualizacion(APP_VERSION)
            if nueva_ver:
                self.after(0, lambda: self._mostrar_update(nueva_ver, url_exe))
            elif manual:
                self.after(0, lambda: messagebox.showinfo(
                    "Sin actualizaciones",
                    "Ya tenes la version mas reciente. Version instalada: v" + APP_VERSION))
        except Exception as e:
            if manual:
                self.after(0, lambda err=str(e): messagebox.showerror(
                    "Error", "No se pudo verificar actualizaciones. " + err))

    def _mostrar_update(self, nueva_ver, url_exe):
        import tempfile
        resp = messagebox.askyesno(
            "Actualizacion disponible",
            "Nueva version v" + str(nueva_ver) + ". Version actual v" + APP_VERSION + ". Queres descargar e instalar ahora?")
        if not resp:
            return
        top = tk.Toplevel(self)
        top.title("Descargando...")
        top.resizable(False, False)
        top.grab_set()
        top.configure(bg=BG)
        top.geometry("380x130+" + str(self.winfo_x()+100) + "+" + str(self.winfo_y()+150))
        tk.Label(top, text="Descargando v" + str(nueva_ver) + "...",
                 font=("Segoe UI", 10, "bold"), bg=BG, fg=AZUL_OSCURO).pack(pady=(16,8))
        canvas = tk.Canvas(top, height=18, bg="#e0e0e0",
                           highlightthickness=1, highlightbackground="#ccc", width=340)
        canvas.pack(padx=20)
        lbl_pct = tk.Label(top, text="0%", font=("Segoe UI", 8), fg="#888", bg=BG)
        lbl_pct.pack()
        dest = os.path.join(tempfile.gettempdir(), "Setup_CargadorHoras_update.exe")
        def progress_cb(pct):
            w = int(340 * pct / 100)
            canvas.delete("all")
            canvas.create_rectangle(0, 0, 340, 18, fill="#e0e0e0", outline="")
            canvas.create_rectangle(0, 0, w, 18, fill=AZUL_CLARO, outline="")
            lbl_pct.config(text=str(int(pct)) + "%")
            top.update_idletasks()
        def do_download():
            ok = descargar_actualizacion(url_exe, dest, progress_cb)
            if ok:
                top.destroy()
                messagebox.showinfo("Listo", "Actualizacion descargada. Se ejecutara el instalador y la app se cerrara.")
                import subprocess
                subprocess.Popen([dest])
                self.destroy()
            else:
                top.destroy()
                messagebox.showerror("Error", "No se pudo descargar la actualizacion.")
        threading.Thread(target=do_download, daemon=True).start()

    def _build(self):
        # Header con logo
        hf = tk.Frame(self, bg=AZUL_OSCURO)
        hf.pack(fill="x")
        # Lado izquierdo blanco para el logo — ancho fijo para no distorsionar
        hf_left = tk.Frame(hf, bg="white", width=340, height=80)
        hf_left.pack(side="left")
        hf_left.pack_propagate(False)
        logo_path = get_asset("logo_app.png")
        if os.path.exists(logo_path):
            try:
                from PIL import Image, ImageTk
                img = Image.open(logo_path).convert("RGB")
                img.thumbnail((320, 72), Image.LANCZOS)
                photo = ImageTk.PhotoImage(img)
                lbl   = tk.Label(hf_left, image=photo, bg="white")
                lbl.image = photo
                lbl.place(relx=0.5, rely=0.5, anchor="center")
            except Exception:
                tk.Label(hf_left, text=APP_AUTHOR, font=("Segoe UI", 13, "bold"),
                         bg="white", fg=AZUL_OSCURO).place(relx=0.5, rely=0.5, anchor="center")
        else:
            tk.Label(hf_left, text=APP_AUTHOR, font=("Segoe UI", 13, "bold"),
                     bg="white", fg=AZUL_OSCURO).place(relx=0.5, rely=0.5, anchor="center")
        # Lado derecho azul con version
        tk.Label(hf, text=f"v{APP_VERSION}", font=("Segoe UI", 9),
                 bg=AZUL_OSCURO, fg="#aaccee").pack(side="right", padx=16, pady=28)

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=10, pady=10)
        self._tab_config(nb)
        self._tab_clientes(nb)
        self._tab_ejecutar(nb)
        self._tab_acerca(nb)
        self._tab_ayuda(nb)

    # --- TAB CONFIGURACIÓN ---
    def _add_context_menu(self, widget, allow_copy=True):
        """Agrega menú contextual de copiar/pegar a un Entry."""
        menu = tk.Menu(widget, tearoff=0)
        if allow_copy:
            menu.add_command(label="Copiar", command=lambda: widget.event_generate("<<Copy>>"))
            menu.add_command(label="Cortar", command=lambda: widget.event_generate("<<Cut>>"))
        menu.add_command(label="Pegar",  command=lambda: widget.event_generate("<<Paste>>"))
        menu.add_separator()
        menu.add_command(label="Seleccionar todo",
                         command=lambda: widget.select_range(0, "end"))
        def show_menu(event):
            try: menu.tk_popup(event.x_root, event.y_root)
            finally: menu.grab_release()
        widget.bind("<Button-3>", show_menu)

    def _tab_config(self, nb):
        f = ttk.Frame(nb, padding=15)
        nb.add(f, text="⚙  Configuración")

        def lbl(t, r, bold=False):
            ttk.Label(f, text=t,
                      font=("Segoe UI", 10, "bold") if bold else ("Segoe UI", 9)
                      ).grid(row=r, column=0, sticky="w", pady=3)

        def entry(var, r, show=""):
            ttk.Entry(f, textvariable=var, width=46, show=show
                      ).grid(row=r, column=1, sticky="ew", padx=(10,0))

        def combo(var, vals, r, w=44):
            ttk.Combobox(f, textvariable=var, values=vals, width=w, state="readonly"
                         ).grid(row=r, column=1, sticky="w", padx=(10,0))

        lbl("Conexión a Redmine", 0, bold=True)
        self.v_url     = tk.StringVar(value=self.cfg["redmine_url"])
        self.v_api_key = tk.StringVar(value=self.cfg["api_key"])
        lbl("URL de Redmine", 1)
        _e_url = ttk.Entry(f, textvariable=self.v_url, width=46)
        _e_url.grid(row=1, column=1, sticky="ew", padx=(10,0))
        self._add_context_menu(_e_url)
        ttk.Label(f, text="API Key").grid(row=2, column=0, sticky="w", pady=3)
        api_frame = ttk.Frame(f)
        api_frame.grid(row=2, column=1, sticky="ew", padx=(10,0))
        _e_api = ttk.Entry(api_frame, textvariable=self.v_api_key, width=38, show="•")
        _e_api.pack(side="left")
        self._add_context_menu(_e_api, allow_copy=False)
        # Bloquear Ctrl+C y Ctrl+X completamente
        _e_api.bind("<<Copy>>", lambda e: "break")
        _e_api.bind("<<Cut>>", lambda e: "break")
        _e_api.bind("<Control-c>", lambda e: "break")
        _e_api.bind("<Control-x>", lambda e: "break")
        ttk.Button(api_frame, text="?", width=2,
                   command=self._ir_a_ayuda).pack(side="left", padx=(4,0))

        ttk.Separator(f).grid(row=3, column=0, columnspan=2, sticky="ew", pady=8)
        lbl("Archivo Excel", 4, bold=True)
        self.v_excel = tk.StringVar(value=self.cfg["archivo_excel"])
        self.v_hoja  = tk.StringVar(value=self.cfg["hoja"])
        lbl("Ruta del Excel", 5)
        xf = ttk.Frame(f)
        xf.grid(row=5, column=1, sticky="ew", padx=(10,0))
        _e_exc = ttk.Entry(xf, textvariable=self.v_excel, width=43)
        _e_exc.pack(side="left")
        self._add_context_menu(_e_exc)
        ttk.Button(xf, text="📂", width=3, command=self._browse_excel).pack(side="left", padx=4)
        lbl("Hoja", 6)
        _e_hoja = ttk.Entry(f, textvariable=self.v_hoja, width=46)
        _e_hoja.grid(row=6, column=1, sticky="ew", padx=(10,0))
        self._add_context_menu(_e_hoja)

        lbl("Excel con columna ID_Ticket", 7)
        self.v_usa_id_ticket = tk.StringVar(value="Si" if self.cfg.get("usa_id_ticket", True) else "No")
        ttk.Combobox(f, textvariable=self.v_usa_id_ticket, values=["Si", "No"],
                     width=8, state="readonly").grid(row=7, column=1, sticky="w", padx=(10,0))

        ttk.Button(f, text="📥  Descargar template Excel",
                   command=self._descargar_template).grid(
            row=8, column=0, columnspan=2, sticky="w", pady=(4,0))

        ttk.Separator(f).grid(row=9, column=0, columnspan=2, sticky="ew", pady=8)
        lbl("Valores por defecto", 10, bold=True)
        self.v_actividad  = tk.StringVar(value=self.cfg["actividad"])
        self.v_tipo_hora  = tk.StringVar(value=self.cfg["tipo_hora"])
        self.v_ubicacion  = tk.StringVar(value=self.cfg["ubicacion"])
        _dia_idx = self.cfg["dia_remoto"]
        self.v_dia_remoto = tk.StringVar(value="(Ninguno)" if _dia_idx < 0 else DIAS_SEMANA[_dia_idx])

        lbl("Actividad",             11);  combo(self.v_actividad, ACTIVIDADES, 11)
        lbl("Tipo de Hora",          12); combo(self.v_tipo_hora, TIPOS_HORA, 12)
        lbl("Ubicación diaria",      13); combo(self.v_ubicacion, UBICACIONES, 13)
        lbl("Día de trabajo remoto", 14); combo(self.v_dia_remoto, ["(Ninguno)"] + DIAS_SEMANA, 14, w=20)
        ttk.Label(f, text="(ese día se cargará como Remoto)",
                  font=("Segoe UI", 8), foreground="gray").grid(
            row=15, column=0, columnspan=2, sticky="w", pady=(0,4))

        ttk.Separator(f).grid(row=16, column=0, columnspan=2, sticky="ew", pady=8)
        # Fila de botones: Guardar + Contraseña juntos
        btn_frame = ttk.Frame(f)
        btn_frame.grid(row=17, column=0, columnspan=2, pady=8)
        ttk.Button(btn_frame, text="💾  Guardar",
                   command=self._guardar_config).pack(side="left", padx=(0, 8))
        ttk.Button(btn_frame, text="🔒  Contraseña",
                   command=self._config_password).pack(side="left")
        f.columnconfigure(1, weight=1)

    def _config_password(self):
        """Popup para configurar o quitar la contraseña de acceso."""
        auth = cargar_auth()
        tiene_pass = bool(auth.get("hash"))

        top = tk.Toplevel(self)
        top.title("Contraseña de acceso")
        top.resizable(False, False)
        top.grab_set()
        top.configure(bg=BG)
        self._center_child(top, 340, 260 if tiene_pass else 220)

        tk.Label(top, text="🔒  Contraseña de acceso",
                 font=("Segoe UI", 11, "bold"), bg=BG, fg=AZUL_OSCURO).pack(pady=(16,8))

        if tiene_pass:
            tk.Label(top, text="Ya tenés una contraseña configurada.",
                     font=("Segoe UI", 9), bg=BG, fg="#555").pack()

        body = tk.Frame(top, bg=BG, padx=20)
        body.pack(fill="x", pady=8)

        if tiene_pass:
            tk.Label(body, text="Contraseña actual:", bg=BG,
                     font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w", pady=4)
            v_actual = tk.StringVar()
            ttk.Entry(body, textvariable=v_actual, show="•", width=24).grid(
                row=0, column=1, padx=(8,0))
        else:
            v_actual = None

        tk.Label(body, text="Nueva contraseña:", bg=BG,
                 font=("Segoe UI", 9)).grid(row=1, column=0, sticky="w", pady=4)
        v_nueva = tk.StringVar()
        ttk.Entry(body, textvariable=v_nueva, show="•", width=24).grid(
            row=1, column=1, padx=(8,0))

        tk.Label(body, text="Confirmar:", bg=BG,
                 font=("Segoe UI", 9)).grid(row=2, column=0, sticky="w", pady=4)
        v_confirm = tk.StringVar()
        ttk.Entry(body, textvariable=v_confirm, show="•", width=24).grid(
            row=2, column=1, padx=(8,0))

        lbl_error = tk.Label(top, text="", fg="#e74c3c", bg=BG, font=("Segoe UI", 8))
        lbl_error.pack()

        bf = tk.Frame(top, bg=BG)
        bf.pack(pady=8)

        def guardar():
            if tiene_pass:
                if hash_pass(v_actual.get()) != auth["hash"]:
                    lbl_error.config(text="Contraseña actual incorrecta.")
                    return
            nueva = v_nueva.get()
            if nueva != v_confirm.get():
                lbl_error.config(text="Las contraseñas no coinciden.")
                return
            if nueva == "":
                guardar_auth({"hash": None})
                messagebox.showinfo("Listo", "Contraseña eliminada. La app no pedirá contraseña.")
            else:
                guardar_auth({"hash": hash_pass(nueva)})
                messagebox.showinfo("Listo", "Contraseña configurada correctamente.")
            top.destroy()

        def quitar():
            if tiene_pass:
                if hash_pass(v_actual.get()) != auth["hash"]:
                    lbl_error.config(text="Contraseña actual incorrecta.")
                    return
            guardar_auth({"hash": None})
            messagebox.showinfo("Listo", "Contraseña eliminada.")
            top.destroy()

        ttk.Button(bf, text="💾  Guardar", command=guardar).pack(side="left", padx=4)
        if tiene_pass:
            ttk.Button(bf, text="🗑  Quitar contraseña", command=quitar).pack(side="left", padx=4)
        ttk.Button(bf, text="Cancelar", command=top.destroy).pack(side="left", padx=4)

    def _center_child(self, win, w, h):
        self.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - w) // 2
        y = self.winfo_y() + (self.winfo_height() - h) // 2
        win.geometry(f"{w}x{h}+{x}+{y}")

    def _ir_a_ayuda(self):
        """Navega a la pestaña de Ayuda."""
        # Buscar el notebook y seleccionar pestaña Ayuda (índice 4)
        for widget in self.winfo_children():
            if isinstance(widget, ttk.Notebook):
                widget.select(4)
                break

    def _browse_excel(self):
        p = filedialog.askopenfilename(
            title="Seleccioná tu Excel",
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("Todos", "*.*")])
        if p: self.v_excel.set(p)

    def _descargar_template(self):
        """Copia el template Excel bundleado según configuración de columnas."""
        usa_id_ticket = self.v_usa_id_ticket.get() == "Si"
        src_name = "Carga_Horas_-_c_ID_Ticket.xlsx" if usa_id_ticket else "Carga_Horas_-_sin_ID_Ticket.xlsx"
        src = get_asset(src_name)
        if not os.path.exists(src):
            messagebox.showwarning("Template no encontrado",
                                   f"No se encontró el archivo {src_name}.")
            return
        dest = filedialog.asksaveasfilename(
            title="Guardar template Excel",
            initialfile="Carga Horas.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")])
        if dest:
            shutil.copy2(src, dest)
            # Actualizar ruta del Excel automáticamente
            self.v_excel.set(dest)
            self.cfg["archivo_excel"] = dest
            guardar_config(self.cfg)
            tipo = "con ID_Ticket" if usa_id_ticket else "sin ID_Ticket"
            messagebox.showinfo("Listo", "Template (" + tipo + ") guardado.\nRuta del Excel actualizada automaticamente.")

    def _guardar_config(self):
        self.cfg.update({
            "redmine_url":      self.v_url.get().strip(),
            "api_key":          self.v_api_key.get().strip(),
            "archivo_excel":    self.v_excel.get().strip(),
            "hoja":             self.v_hoja.get().strip(),
            "actividad":        self.v_actividad.get(),
            "tipo_hora":        self.v_tipo_hora.get(),
            "ubicacion":        self.v_ubicacion.get(),
            "ubicacion_remoto": "Remoto",
            "dia_remoto":       -1 if self.v_dia_remoto.get() == "(Ninguno)" else DIAS_SEMANA.index(self.v_dia_remoto.get()),
            "usa_id_ticket":         self.v_usa_id_ticket.get() == "Si",
        })
        guardar_config(self.cfg)
        # Actualizar URL en pestaña Acerca
        if hasattr(self, "_lbl_redmine_url"):
            self._lbl_redmine_url.config(text=self.cfg.get("redmine_url", ""))
        messagebox.showinfo("Guardado", "✔ Configuración guardada.")

    # --- TAB CLIENTES ---
    def _tab_clientes(self, nb):
        f = ttk.Frame(nb, padding=15)
        nb.add(f, text="🏢  Clientes")

        ttk.Label(f, text="Conectate a Redmine para ver los proyectos disponibles.\n"
                          "Tildá los que usás y asignales el nombre que usás en tu Excel. Podés agregar más con +.",
                  font=("Segoe UI", 9), foreground="gray").pack(anchor="w")

        btn_frame = ttk.Frame(f)
        btn_frame.pack(anchor="w", pady=(8,4))
        ttk.Button(btn_frame, text="🔄  Cargar proyectos desde Redmine",
                   command=self._cargar_proyectos).pack(side="left")
        ttk.Button(btn_frame, text="🔍  Solo tildados",
                   command=self._filtrar_tildados).pack(side="left", padx=(8,0))

        self.lbl_cli = ttk.Label(f, text="", foreground="gray", font=("Segoe UI", 8))
        self.lbl_cli.pack(anchor="w")

        # Encabezados de tabla
        hdr = tk.Frame(f, bg=AZUL_OSCURO)
        hdr.pack(fill="x", pady=(6,0))
        for txt, w in [("✔", 3), ("Proyecto en Redmine", 22), ("Nombre en Excel", 38)]:
            tk.Label(hdr, text=txt, width=w, bg=AZUL_OSCURO, fg="white",
                     font=("Segoe UI", 9, "bold"), anchor="w").pack(side="left", padx=4)

        # Canvas scrollable dentro de un frame para no interferir con el botón
        table_frame = tk.Frame(f)
        table_frame.pack(fill="both", expand=True)
        canvas = tk.Canvas(table_frame, height=240, bg="white", highlightthickness=0)
        self._cli_canvas = canvas  # store ref for scroll-to-top
        sb     = ttk.Scrollbar(table_frame, orient="vertical", command=canvas.yview)
        self.cli_inner = tk.Frame(canvas, bg="white")
        self.cli_inner.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0), window=self.cli_inner, anchor="nw")
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        # Scroll con rueda del mouse - bind recursivo para todos los widgets hijos
        def _on_mousewheel(event):
            # Solo scrollear si el contenido supera la altura del canvas
            canvas.update_idletasks()
            content_h = self.cli_inner.winfo_reqheight()
            canvas_h  = canvas.winfo_height()
            if content_h > canvas_h:
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        def _bind_mousewheel(widget):
            widget.bind("<MouseWheel>", _on_mousewheel)
            for child in widget.winfo_children():
                _bind_mousewheel(child)
        canvas.bind("<MouseWheel>", _on_mousewheel)
        self.cli_inner.bind("<MouseWheel>", _on_mousewheel)
        self._cli_bind_scroll = _bind_mousewheel

        # Precargar guardados
        for c in self.clientes:
            nombres = c.get("nombres_excel", c.get("nombre_excel", ""))
            self._fila_cliente({"id": c["proyecto_id"], "name": c["proyecto_nombre"]},
                               True, nombres)

        if self.clientes:
            self.lbl_cli.config(text=f"✔ {len(self.clientes)} clientes guardados.")
        # Actualizar scrollregion después de precargar
        self.cli_inner.update_idletasks()
        if hasattr(self, "_cli_canvas"):
            self._cli_canvas.configure(scrollregion=self._cli_canvas.bbox("all"))
            self._cli_canvas.yview_moveto(0)

        # Botón FUERA del scroll, siempre visible abajo
        ttk.Separator(f).pack(fill="x", pady=(8,0))
        ttk.Button(f, text="💾  Guardar configuración de clientes",
                   command=self._guardar_clientes).pack(anchor="center", pady=(8,0))

    def _fila_cliente(self, proyecto, tildado=False, nombres_excel=""):
        """Crea un bloque de filas para un proyecto. Cada nombre va en su propia línea."""
        bloque = tk.Frame(self.cli_inner, bg="white", bd=0)
        bloque.pack(fill="x", pady=1)

        vc = tk.BooleanVar(value=tildado)
        nombres = [n.strip() for n in nombres_excel.split(",") if n.strip()] if nombres_excel else []
        if not nombres:
            nombres = [proyecto["name"]]

        vars_nombre = []
        filas_frame = []

        def agregar_fila_nombre(valor=""):
            fila = tk.Frame(bloque, bg="white")
            fila.pack(fill="x")
            es_primera = len(vars_nombre) == 0
            if es_primera:
                ttk.Checkbutton(fila, variable=vc).pack(side="left", padx=4)
                tk.Label(fila, text=proyecto["name"][:24], width=22, bg="white",
                         anchor="w", font=("Segoe UI", 9)).pack(side="left")
            else:
                tk.Label(fila, text="", width=26, bg="white").pack(side="left")
            vn = tk.StringVar(value=valor)
            ttk.Entry(fila, textvariable=vn, width=26).pack(side="left", padx=4)
            ttk.Button(fila, text="+", width=2,
                       command=lambda: agregar_fila_nombre("")).pack(side="left")
            # Botón eliminar solo para filas extra (no la primera)
            if not es_primera:
                def eliminar(f=fila, v=vn):
                    idx = vars_nombre.index(v)
                    vars_nombre.pop(idx)
                    filas_frame.pop(idx)
                    f.destroy()
                ttk.Button(fila, text="✕", width=2,
                           command=eliminar).pack(side="left", padx=(2,0))
            vars_nombre.append(vn)
            filas_frame.append(fila)
            # Re-bindear scroll
            if hasattr(self, "_cli_bind_scroll"):
                self._cli_bind_scroll(bloque)

        for nom in nombres:
            agregar_fila_nombre(nom)

        self._cli_rows.append((vc, vars_nombre, proyecto))

    def _filtrar_tildados(self):
        """Oculta las filas no tildadas, muestra solo las activas."""
        for w in self.cli_inner.winfo_children():
            w.destroy()
        guardados_rows = [(vc, vn, p) for vc, vn, p in self._cli_rows if vc.get()]
        self._cli_rows.clear()
        for vc_old, vn_list, p in guardados_rows:
            nombres = ",".join(v.get() for v in vn_list if v.get().strip())
            self._fila_cliente(p, True, nombres)
        self.lbl_cli.config(text=f"Mostrando {len(self._cli_rows)} cliente(s) tildado(s).")
        if hasattr(self, "_cli_bind_scroll"):
            self._cli_bind_scroll(self.cli_inner)
        # Resetear scroll region y volver al inicio
        if hasattr(self, "_cli_canvas"):
            self.cli_inner.update_idletasks()
            self._cli_canvas.configure(scrollregion=self._cli_canvas.bbox("all"))
            self._cli_canvas.yview_moveto(0)

    def _cargar_proyectos(self):
        url = self.cfg.get("redmine_url","")
        key = self.cfg.get("api_key","")
        if not key:
            messagebox.showwarning("Falta API Key", "Guardá tu API Key en Configuración primero.")
            return
        self.lbl_cli.config(text="⏳ Cargando proyectos...")
        self.update_idletasks()

        def fetch():
            proyectos, err = obtener_proyectos(url, key)
            if err:
                self.lbl_cli.config(text=f"❌ Error: {err}"); return
            for w in self.cli_inner.winfo_children(): w.destroy()
            self._cli_rows.clear()
            # Mapeo proyecto_id → nombres guardados
            guardados = {c["proyecto_id"]: c.get("nombres_excel", c.get("nombre_excel",""))
                         for c in self.clientes}
            for p in sorted(proyectos, key=lambda x: x["name"]):
                self._fila_cliente(p, p["id"] in guardados, guardados.get(p["id"], ""))
            self.lbl_cli.config(text=f"✔ {len(proyectos)} proyectos cargados.")
            # Re-bindear scroll y actualizar region
            self.cli_inner.update_idletasks()
            if hasattr(self, "_cli_canvas"):
                self._cli_canvas.configure(scrollregion=self._cli_canvas.bbox("all"))
            if hasattr(self, "_cli_bind_scroll"):
                self._cli_bind_scroll(self.cli_inner)
        threading.Thread(target=fetch, daemon=True).start()

    def _aplicar_validacion_clientes(self, ruta_excel, clientes_lista):
        """Escribe una hoja oculta _Clientes con los nombres validos
        y aplica DataValidation a la columna Cliente de la hoja principal."""
        if not ruta_excel or not os.path.exists(ruta_excel):
            return False, "El archivo Excel no existe"
        try:
            import openpyxl
            from openpyxl.worksheet.datavalidation import DataValidation
            wb = openpyxl.load_workbook(ruta_excel)

            # Reunir todos los nombres validos (case-insensitive ya en la app)
            nombres = []
            for c in clientes_lista:
                for n in c.get("nombres_excel", "").split(","):
                    n = n.strip()
                    if n and n not in nombres:
                        nombres.append(n)

            # Agregar tambien los clientes especiales
            for esp in ["Licencia", "Vacaciones"]:
                if esp not in nombres:
                    nombres.append(esp)

            if not nombres:
                return False, "No hay clientes para validar"

            # Crear o actualizar la hoja _Clientes
            sheet_name = "_Clientes"
            if sheet_name in wb.sheetnames:
                ws_cli = wb[sheet_name]
                # Limpiar contenido anterior
                ws_cli.delete_rows(1, ws_cli.max_row)
            else:
                ws_cli = wb.create_sheet(sheet_name)

            # Escribir nombres en columna A
            for i, nombre in enumerate(nombres, 1):
                ws_cli.cell(row=i, column=1, value=nombre)

            # Ocultar la hoja
            ws_cli.sheet_state = "hidden"

            # Aplicar validacion en la hoja principal
            hoja_principal = self.cfg.get("hoja", "Horas")
            if hoja_principal not in wb.sheetnames:
                wb.save(ruta_excel)
                return False, f"No se encontro la hoja \"{hoja_principal}\""

            ws_main = wb[hoja_principal]

            # Buscar columna Cliente
            header_row = [cell.value for cell in ws_main[1]]
            try:
                col_cliente = header_row.index("Cliente") + 1
            except ValueError:
                wb.save(ruta_excel)
                return False, "No se encontro la columna \"Cliente\""

            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col_cliente)

            # Eliminar validaciones previas en esa columna
            keep = []
            for dv in ws_main.data_validations.dataValidation:
                if not any(col_letter in str(r) for r in dv.sqref.ranges):
                    keep.append(dv)
            ws_main.data_validations.dataValidation = keep

            # Crear nueva validacion
            rango_lista = "$A$1:$A$" + str(len(nombres))
            formula = "=" + sheet_name + "!" + rango_lista
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            dv.error = "Cliente no valido. Usa uno de la lista."
            dv.errorTitle = "Cliente invalido"
            dv.prompt = "Seleccionar de la lista"
            dv.promptTitle = "Cliente"
            dv.showErrorMessage = False  # No bloquear, solo sugerir
            ws_main.add_data_validation(dv)

            # Aplicar al rango de la columna (filas 2 a 1000)
            dv.add(col_letter + "2:" + col_letter + "1000")

            wb.save(ruta_excel)
            return True, f"Validacion aplicada a {len(nombres)} cliente(s)"
        except Exception as e:
            return False, str(e)

    def _excel_esta_abierto(self, ruta):
        """Detecta si el archivo Excel esta abierto/bloqueado por otro proceso."""
        if not ruta or not os.path.exists(ruta):
            return False
        try:
            # Intentar abrir en modo append (requiere acceso exclusivo)
            with open(ruta, "a"):
                pass
            return False
        except (PermissionError, OSError):
            return True

    def _aviso_excel_abierto(self, ruta):
        """Muestra un mensaje claro indicando que el Excel debe cerrarse."""
        nombre = os.path.basename(ruta) if ruta else "el archivo Excel"
        messagebox.showwarning(
            "Excel abierto",
            "El archivo \"" + nombre + "\" esta abierto en Excel.\n\n"
            "Cerralo y volve a intentar la operacion.")

    def _guardar_clientes(self):
        # Verificar que el Excel no este abierto (en caso futuro de aplicar validacion)
        ruta = self.cfg.get("archivo_excel", "")
        if ruta and self._excel_esta_abierto(ruta):
            self._aviso_excel_abierto(ruta)
            return
        res = []
        for vc, vars_nombre, p in self._cli_rows:
            if vc.get():
                nombres = [v.get().strip() for v in vars_nombre if v.get().strip()]
                if nombres:
                    res.append({
                        "nombres_excel":   ",".join(nombres),
                        "proyecto_id":     p["id"],
                        "proyecto_nombre": p["name"]
                    })
        self.clientes = res
        guardar_clientes(res)
        total = sum(len(c["nombres_excel"].split(",")) for c in res)

        # Aplicar validacion de datos al Excel si esta configurado
        msg_validacion = ""
        if ruta and os.path.exists(ruta) and res:
            ok_val, info = self._aplicar_validacion_clientes(ruta, res)
            if ok_val:
                msg_validacion = "\n\n✔ Lista desplegable aplicada al Excel."
            else:
                msg_validacion = f"\n\n⚠ No se pudo aplicar al Excel: {info}"

        messagebox.showinfo("Guardado",
            f"✔ {len(res)} proyecto(s), {total} nombre(s) guardados." + msg_validacion)

    # --- TAB EJECUTAR ---
    def _tab_ejecutar(self, nb):
        f = ttk.Frame(nb, padding=15)
        nb.add(f, text="▶  Ejecutar carga")

        # Modo de carga
        ttk.Label(f, text="Modo de carga",
                  font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0,4))

        modo_frame = ttk.Frame(f)
        modo_frame.pack(fill="x", pady=(0,6))

        self.v_modo = tk.StringVar(value="completo")

        rb1 = ttk.Radiobutton(modo_frame, text="Cargar Excel completo",
                              variable=self.v_modo, value="completo",
                              command=self._toggle_fechas)
        rb1.pack(side="left")
        rb2 = ttk.Radiobutton(modo_frame, text="Selección de fechas",
                              variable=self.v_modo, value="fechas",
                              command=self._toggle_fechas)
        rb2.pack(side="left", padx=(20,0))

        # Frame de fechas (oculto por defecto)
        self.fechas_frame = ttk.Frame(f)
        self.fechas_frame.pack(fill="x", pady=(0,6))

        ttk.Label(self.fechas_frame, text="Desde:").pack(side="left")
        self.v_fecha_desde = tk.StringVar()
        ttk.Entry(self.fechas_frame, textvariable=self.v_fecha_desde,
                  width=12, state="readonly").pack(side="left", padx=(4,4))
        ttk.Button(self.fechas_frame, text="📅", width=3,
                   command=lambda: self._abrir_calendario(self.v_fecha_desde)
                   ).pack(side="left", padx=(0,16))

        ttk.Label(self.fechas_frame, text="Hasta:").pack(side="left")
        self.v_fecha_hasta = tk.StringVar()
        ttk.Entry(self.fechas_frame, textvariable=self.v_fecha_hasta,
                  width=12, state="readonly").pack(side="left", padx=(4,4))
        ttk.Button(self.fechas_frame, text="📅", width=3,
                   command=lambda: self._abrir_calendario(self.v_fecha_hasta)
                   ).pack(side="left")

        self.fechas_frame.pack_forget()  # oculto al inicio

        ttk.Separator(f).pack(fill="x", pady=(4,8))

        # Panel de resultados amigable — siempre visible
        self.resultado_frame = tk.Frame(f, bg="#f0f4f8", relief="solid", bd=1)
        self.resultado_frame.pack(fill="x", pady=(0,8))

        self._lbl_res = {}
        res_inner = tk.Frame(self.resultado_frame, bg="#f0f4f8", padx=8, pady=8)
        res_inner.pack(fill="x")
        for key, linea1, linea2, color in [
            ("issues",  "Issues",    "creados",   AZUL_OSCURO),
            ("ok",      "Horas",     "cargadas",  "#2ecc71"),
            ("dup",     "Duplic.",   "omitidos",  "#f39c12"),
            ("errores", "Con",       "errores",   "#e74c3c"),
            ("omitidas","Filas",     "omitidas",  "#888888"),
        ]:
            rf = tk.Frame(res_inner, bg="#f0f4f8")
            rf.pack(side="left", expand=True)
            tk.Label(rf, text=linea1, font=("Segoe UI", 8, "bold"),
                     bg="#f0f4f8", fg=color).pack()
            tk.Label(rf, text=linea2, font=("Segoe UI", 8, "bold"),
                     bg="#f0f4f8", fg=color).pack()
            lbl = tk.Label(rf, text="0", font=("Segoe UI", 20, "bold"),
                           bg="#f0f4f8", fg=color)
            lbl.pack()
            self._lbl_res[key] = lbl

        # Log colapsable
        self._log_expanded = tk.BooleanVar(value=False)
        log_hdr = tk.Frame(f, bg=BG)
        log_hdr.pack(fill="x")
        self._log_toggle_lbl = tk.Label(log_hdr, text="+ Ver log detallado",
                                         font=("Segoe UI", 9), fg=AZUL_CLARO,
                                         bg=BG, cursor="hand2")
        self._log_toggle_lbl.pack(side="left")
        ttk.Button(log_hdr, text="🗑", width=3,
                   command=self._limpiar_log).pack(side="right")
        ttk.Button(log_hdr, text="📋 Log", width=6,
                   command=self._abrir_log).pack(side="right", padx=(0,4))
        self._log_toggle_lbl.bind("<Button-1>", self._toggle_log_ejecutar)

        self._log_frame_ejecutar = tk.Frame(f, bg=BG)
        self.log_box = scrolledtext.ScrolledText(
            self._log_frame_ejecutar, width=72, height=14, state="disabled",
            font=("Consolas", 8), bg="#1e1e1e", fg="#d4d4d4")
        self.log_box.pack(fill="both", expand=True)

        bf = ttk.Frame(f)
        bf.pack(fill="x", pady=(8,0))
        self.btn_run = ttk.Button(bf, text="▶  Iniciar carga", command=self._iniciar)
        self.btn_run.pack(side="left")

        # Barra de progreso personalizada estilo Windows
        self._progress_canvas = tk.Canvas(f, height=18, bg="#e0e0e0",
                                           highlightthickness=1,
                                           highlightbackground="#cccccc")
        self._progress_canvas.pack(fill="x", pady=(6,0))
        self._progress_rect  = None
        self._progress_anim  = False
        self._progress_pos   = 0
        self.progress_lbl = tk.Label(f, text="Listo para iniciar",
                                      font=("Segoe UI", 8), fg="#888", bg=BG)
        self.progress_lbl.pack(anchor="w")

    def _abrir_calendario(self, var):
        """Abre un popup de calendario para seleccionar fecha."""
        from datetime import date
        top = tk.Toplevel(self)
        top.title("Seleccionar fecha")
        top.resizable(False, False)
        top.grab_set()

        # Centrar sobre la ventana principal
        top.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() // 2) - 150
        y = self.winfo_y() + (self.winfo_height() // 2) - 130
        top.geometry(f"300x260+{x}+{y}")

        from datetime import datetime
        try:
            current = datetime.strptime(var.get(), "%d/%m/%Y")
        except ValueError:
            current = date.today()
            current = datetime(current.year, current.month, current.day)

        self._cal_year  = tk.IntVar(value=current.year)
        self._cal_month = tk.IntVar(value=current.month)

        MESES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
                 "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

        header = tk.Frame(top, bg=AZUL_OSCURO)
        header.pack(fill="x")
        tk.Button(header, text="◀", bg=AZUL_OSCURO, fg="white", bd=0, font=("Segoe UI",10),
                  command=lambda: self._cal_nav(-1, cal_frame, var, top)
                  ).pack(side="left", padx=8, pady=4)
        self._cal_lbl = tk.Label(header, bg=AZUL_OSCURO, fg="white",
                                  font=("Segoe UI", 10, "bold"))
        self._cal_lbl.pack(side="left", expand=True)
        tk.Button(header, text="▶", bg=AZUL_OSCURO, fg="white", bd=0, font=("Segoe UI",10),
                  command=lambda: self._cal_nav(1, cal_frame, var, top)
                  ).pack(side="right", padx=8, pady=4)

        cal_frame = tk.Frame(top, bg="white")
        cal_frame.pack(fill="both", expand=True, padx=8, pady=4)

        self._cal_render(cal_frame, var, top)

    def _cal_nav(self, delta, frame, var, top):
        m = self._cal_month.get() + delta
        y = self._cal_year.get()
        if m > 12: m = 1;  y += 1
        if m < 1:  m = 12; y -= 1
        self._cal_month.set(m)
        self._cal_year.set(y)
        self._cal_render(frame, var, top)

    def _cal_render(self, frame, var, top):
        import calendar
        from datetime import date
        MESES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
                 "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
        for w in frame.winfo_children():
            w.destroy()
        y = self._cal_year.get()
        m = self._cal_month.get()
        self._cal_lbl.config(text=f"{MESES[m-1]} {y}")
        dias = ["Lu","Ma","Mi","Ju","Vi","Sa","Do"]
        for col, d in enumerate(dias):
            tk.Label(frame, text=d, font=("Segoe UI",8,"bold"),
                     bg=AZUL_OSCURO, fg="white", width=3).grid(row=0, column=col, padx=1, pady=1)
        cal = calendar.monthcalendar(y, m)
        for row, week in enumerate(cal, 1):
            for col, day in enumerate(week):
                if day == 0:
                    tk.Label(frame, text="", bg="white", width=3).grid(row=row, column=col)
                else:
                    def on_click(d=day, y=y, m=m):
                        var.set(f"{d:02d}/{m:02d}/{y}")
                        top.destroy()
                    bg = AZUL_CLARO if col >= 5 else "white"
                    tk.Button(frame, text=str(day), width=3,
                              font=("Segoe UI",9), bg=bg, relief="flat",
                              command=on_click).grid(row=row, column=col, padx=1, pady=1)

    def _toggle_fechas(self):
        if self.v_modo.get() == "fechas":
            self.fechas_frame.pack(fill="x", pady=(0,6),
                                   before=self.fechas_frame.master.winfo_children()[3])
        else:
            self.fechas_frame.pack_forget()

    def _log(self, msg):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", msg+"\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")
        self.update_idletasks()

    def _mostrar_resultados(self):
        """Parsea el log y actualiza el panel de resultados."""
        import re
        log_text = self.log_box.get("1.0", "end")
        def extraer(patron):
            m = re.search(patron, log_text)
            return m.group(1) if m else "0"
        self._lbl_res["issues"].config(text=extraer(r"Issues creados\s*:\s*(\d+)"))
        self._lbl_res["ok"].config(text=extraer(r"Horas cargadas\s*:\s*(\d+)"))
        self._lbl_res["dup"].config(text=extraer(r"Duplicados omit\.\s*:\s*(\d+)"))
        self._lbl_res["errores"].config(text=extraer(r"Con errores\s*:\s*(\d+)"))
        self._lbl_res["omitidas"].config(text=extraer(r"Omitidas\s*:\s*(\d+)"))
        # Si hay errores, abrir el log automáticamente
        if int(extraer(r"Con errores\s*:\s*(\d+)")) > 0:
            if not self._log_expanded.get():
                self._toggle_log_ejecutar()

    def _animar_barra(self):
        """Anima la barra estilo Windows — bloque que avanza de izquierda a derecha."""
        if not self._progress_anim:
            return
        c = self._progress_canvas
        c.update_idletasks()
        w = c.winfo_width()
        if w < 2:
            self.after(50, self._animar_barra)
            return
        c.delete("all")
        # Fondo
        c.create_rectangle(0, 0, w, 18, fill="#e0e0e0", outline="")
        # Bloque deslizante (30% del ancho)
        block = int(w * 0.30)
        x1 = int(self._progress_pos) % (w + block) - block
        x2 = x1 + block
        # Gradiente simulado con dos rectángulos
        c.create_rectangle(x1, 0, x2, 18, fill=AZUL_CLARO, outline="")
        c.create_rectangle(x1, 0, x1 + block // 3, 18, fill="#55ccee", outline="")
        self._progress_pos += 8
        self.after(30, self._animar_barra)

    def _toggle_log_ejecutar(self, event=None):
        if self._log_expanded.get():
            self._log_frame_ejecutar.pack_forget()
            self._log_toggle_lbl.config(text="+ Ver log detallado")
            self._log_expanded.set(False)
        else:
            self._log_frame_ejecutar.pack(fill="both", expand=True, pady=(4,0))
            self._log_toggle_lbl.config(text="- Ocultar log detallado")
            self._log_expanded.set(True)

    def _limpiar_log(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0","end")
        self.log_box.configure(state="disabled")
        # Resetear barra
        self._progress_canvas.delete("all")
        self._progress_canvas.create_rectangle(0, 0, 0, 18, fill="#e0e0e0", outline="")
        self.progress_lbl.config(text="Listo para iniciar", fg="#888")
        for lbl in self._lbl_res.values():
            lbl.config(text="0")

    def _iniciar(self):
        if not self.cfg.get("api_key"):
            messagebox.showwarning("Falta configuración","Completá la API Key."); return
        if not self.cfg.get("archivo_excel"):
            messagebox.showwarning("Falta configuración","Seleccioná el Excel."); return
        if not self.clientes:
            messagebox.showwarning("Sin clientes","Configurá clientes en la pestaña Clientes."); return
        # Verificar que el Excel no este abierto
        ruta_excel = self.cfg.get("archivo_excel", "")
        if self._excel_esta_abierto(ruta_excel):
            self._aviso_excel_abierto(ruta_excel)
            return
        # Validar fechas si corresponde
        fecha_desde = fecha_hasta = None
        if self.v_modo.get() == "fechas":
            from datetime import datetime
            try:
                fecha_desde = datetime.strptime(self.v_fecha_desde.get().strip(), "%d/%m/%Y")
                fecha_hasta = datetime.strptime(self.v_fecha_hasta.get().strip(), "%d/%m/%Y")
                if fecha_desde > fecha_hasta:
                    messagebox.showwarning("Fechas inválidas",
                                           "La fecha Desde no puede ser mayor que Hasta.")
                    return
            except ValueError:
                messagebox.showwarning("Fechas inválidas",
                                       "Formato incorrecto. Usá dd/mm/yyyy.")
                return

        self.btn_run.configure(state="disabled")
        self.progress_lbl.config(text="Cargando...", fg=AZUL_CLARO)
        for lbl in self._lbl_res.values():
            lbl.config(text="0")
        self._progress_anim = True
        self._progress_pos  = 0
        self._animar_barra()

        def on_done(ok, horas=0, issues=0):
            self.btn_run.configure(state="normal")
            self._progress_anim = False
            # Mostrar barra completa o vacía
            c = self._progress_canvas
            c.delete("all")
            w = c.winfo_width()
            if ok:
                c.create_rectangle(0, 0, w, 18, fill="#2ecc71", outline="")
            else:
                c.create_rectangle(0, 0, 0, 18, fill="#e74c3c", outline="")
            if ok:
                self.progress_lbl.config(text="✔  Carga finalizada", fg="#2ecc71")
            else:
                self.progress_lbl.config(text="❌  Finalizado con errores — revisá el log", fg="#e74c3c")
            if ok:
                self.stats["horas_cargadas"] += horas
                self.stats["issues_creados"] += issues
                self.stats["sesiones"] += 1
                guardar_stats(self.stats)
                self._actualizar_stats()
            self._mostrar_resultados()
        threading.Thread(target=ejecutar_carga,
                         args=(self.cfg, self.clientes, self._log, on_done,
                               fecha_desde, fecha_hasta),
                         daemon=True).start()

    # --- TAB ACERCA ---
    def _tab_acerca(self, nb):
        f = tk.Frame(nb, bg=BG, padx=20, pady=20)
        nb.add(f, text="ℹ  Acerca")

        # Logo
        icono_path = get_asset("icono_acerca.png")
        if os.path.exists(icono_path):
            try:
                from PIL import Image, ImageTk
                img = Image.open(icono_path).convert("RGBA")
                img.thumbnail((120, 120), Image.LANCZOS)
                photo = ImageTk.PhotoImage(img)
                lbl   = tk.Label(f, image=photo, bg=BG, bd=0)
                lbl.image = photo
                lbl.pack(pady=(0, 8))
            except Exception:
                pass

        # Info app
        tk.Frame(f, bg=AZUL_OSCURO, height=2).pack(fill="x", pady=(0,12))

        info_frame = tk.Frame(f, bg=BG)
        info_frame.pack(fill="x")

        def fila_info(label, valor, color="#333333"):
            rf = tk.Frame(info_frame, bg=BG)
            rf.pack(fill="x", pady=2)
            tk.Label(rf, text=label, font=("Segoe UI", 9, "bold"),
                     bg=BG, fg="#555555", width=18, anchor="w").pack(side="left")
            tk.Label(rf, text=valor, font=("Segoe UI", 9),
                     bg=BG, fg=color).pack(side="left")

        fila_info("Aplicación:", APP_NAME)
        fila_info("Versión:", f"v{APP_VERSION}", AZUL_CLARO)
        fila_info("Desarrollado por:", "Bruno Pérez", AZUL_OSCURO)
        fila_info("Empresa:", APP_AUTHOR)
        rf_rd = tk.Frame(info_frame, bg=BG)
        rf_rd.pack(fill="x", pady=2)
        tk.Label(rf_rd, text="Redmine:", font=("Segoe UI", 9, "bold"),
                 bg=BG, fg="#555555", width=18, anchor="w").pack(side="left")
        lbl_rd = tk.Label(rf_rd, text=self.cfg.get("redmine_url", "-"),
                          font=("Segoe UI", 9), bg=BG, fg="#333333")
        lbl_rd.pack(side="left")
        self._lbl_redmine_url = lbl_rd

        tk.Frame(f, bg="#e0e0e0", height=1).pack(fill="x", pady=12)

        # Estadísticas
        tk.Label(f, text="Estadísticas acumuladas",
                 font=("Segoe UI", 10, "bold"), fg=AZUL_OSCURO,
                 bg=BG).pack(anchor="w")

        stats_frame = tk.Frame(f, bg=BG, pady=8)
        stats_frame.pack(fill="x")

        self.lbl_stat_horas   = tk.Label(stats_frame, font=("Segoe UI", 9),
                                          bg=BG)
        self.lbl_stat_issues  = tk.Label(stats_frame, font=("Segoe UI", 9),
                                          bg=BG)
        self.lbl_stat_sesiones = tk.Label(stats_frame, font=("Segoe UI", 9),
                                           bg=BG)
        self.lbl_stat_horas.pack(anchor="w")
        self.lbl_stat_issues.pack(anchor="w")
        self.lbl_stat_sesiones.pack(anchor="w")
        self._actualizar_stats()

        tk.Frame(f, bg="#e0e0e0", height=1).pack(fill="x", pady=12)
        tk.Frame(f, bg="#e0e0e0", height=1).pack(fill="x", pady=8)
        btn_frame = tk.Frame(f, bg=BG)
        btn_frame.pack()
        ttk.Button(btn_frame, text="🔄  Buscar actualizaciones",
                   command=lambda: threading.Thread(
                       target=lambda: self._check_update(manual=True),
                       daemon=True).start()).pack(side="left", padx=4)
        tk.Label(f, text=f"© {APP_AUTHOR} — Todos los derechos reservados",
                 font=("Segoe UI", 8), fg="gray",
                 bg=BG).pack(pady=(8,0))

    def _actualizar_stats(self):
        s = self.stats
        if hasattr(self, "lbl_stat_horas"):
            self.lbl_stat_horas.config(
                text=f"  ⏱  Horas cargadas totales:   {s.get('horas_cargadas', 0)}")
            self.lbl_stat_issues.config(
                text=f"  📋  Issues creados totales:   {s.get('issues_creados', 0)}")
            self.lbl_stat_sesiones.config(
                text=f"  🔄  Sesiones de carga:         {s.get('sesiones', 0)}")

    def _tab_ayuda(self, nb):
        f = tk.Frame(nb, bg=BG)
        nb.add(f, text="❓  Ayuda")

        canvas = tk.Canvas(f, bg=BG, highlightthickness=0)
        sb = ttk.Scrollbar(f, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, bg=BG, padx=16, pady=12)
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        def _scroll(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")

        def _bind_scroll(w):
            w.bind("<MouseWheel>", _scroll)
            for c in w.winfo_children():
                _bind_scroll(c)

        inner.bind("<Configure>", lambda e: (
            canvas.configure(scrollregion=canvas.bbox("all")),
            _bind_scroll(inner)))
        canvas.bind("<MouseWheel>", _scroll)

        # ── helpers ──────────────────────────────────────────
        def seccion(texto):
            """Título de sección con línea de color."""
            tk.Label(inner, text=texto, font=("Segoe UI", 11, "bold"),
                     bg=BG, fg=AZUL_OSCURO).pack(anchor="w", pady=(18, 2))
            tk.Frame(inner, bg=AZUL_CLARO, height=2).pack(fill="x", pady=(0, 6))

        def subtit(texto):
            tk.Label(inner, text=texto, font=("Segoe UI", 10, "bold"),
                     bg=BG, fg="#2a2a2a").pack(anchor="w", pady=(10, 2))

        def parrafo(texto):
            """Texto seleccionable y copiable."""
            chars = 78
            lines_needed = sum(max(1, (len(p)+chars-1)//chars) for p in texto.split("\n"))
            t = tk.Text(inner, font=("Segoe UI", 9), fg="#333", bg=BG,
                        relief="flat", bd=0, wrap="word", height=max(1, lines_needed),
                        padx=0, pady=1, insertwidth=0)
            t.insert("1.0", texto)
            t.bind("<Key>", lambda e: "break")
            t.bind("<MouseWheel>", _scroll)
            ctx = tk.Menu(t, tearoff=0)
            ctx.add_command(label="Copiar", command=lambda: t.event_generate("<<Copy>>"))
            ctx.add_command(label="Seleccionar todo",
                            command=lambda: t.tag_add("sel", "1.0", "end"))
            t.bind("<Button-3>", lambda e: ctx.tk_popup(e.x_root, e.y_root))
            t.pack(anchor="w", fill="x", pady=1)

        def box(titulo_box, desc_box, color_bg="#f0f4f8", color_titulo=AZUL_OSCURO):
            """Box con título y descripción seleccionable."""
            rf = tk.Frame(inner, bg=color_bg, padx=12, pady=8,
                          relief="solid", bd=1)
            rf.pack(fill="x", pady=3)
            tk.Label(rf, text=titulo_box, font=("Segoe UI", 9, "bold"),
                     bg=color_bg, fg=color_titulo).pack(anchor="w")
            chars = 72
            lines_needed = max(1, (len(desc_box)+chars-1)//chars)
            t = tk.Text(rf, font=("Segoe UI", 9), fg="#444", bg=color_bg,
                        relief="flat", bd=0, wrap="word", height=lines_needed,
                        padx=0, pady=2, insertwidth=0)
            t.insert("1.0", desc_box)
            t.bind("<Key>", lambda e: "break")
            t.bind("<MouseWheel>", _scroll)
            ctx = tk.Menu(t, tearoff=0)
            ctx.add_command(label="Copiar", command=lambda: t.event_generate("<<Copy>>"))
            ctx.add_command(label="Seleccionar todo",
                            command=lambda: t.tag_add("sel", "1.0", "end"))
            t.bind("<Button-3>", lambda e: ctx.tk_popup(e.x_root, e.y_root))
            t.pack(anchor="w", fill="x")

        rdm_url = self.cfg.get("redmine_url", "https://rdm.hmconsulting.com.ar")
        account_url = rdm_url.rstrip("/") + "/my/account"

        # ── API KEY ───────────────────────────────────────────
        seccion("Como obtener tu API Key de Redmine")
        parrafo("La API Key es una clave personal que identifica tu usuario. La app la usa para cargar horas en tu nombre sin necesitar tu contrasena.")
        subtit("Pasos:")
        parrafo("1. Ingresa a tu cuenta en Redmine (hace click en el link):")

        lnk_frame = tk.Frame(inner, bg=BG)
        lnk_frame.pack(anchor="w", pady=2)
        lnk_frame.bind("<MouseWheel>", _scroll)
        lnk = tk.Label(lnk_frame, text="   " + account_url,
                       font=("Segoe UI", 9, "underline"), fg=AZUL_CLARO,
                       bg=BG, cursor="hand2")
        lnk.pack(side="left")
        lnk.bind("<Button-1>", lambda e: __import__("webbrowser").open(account_url))
        lnk.bind("<MouseWheel>", _scroll)
        copy_lnk = tk.Label(lnk_frame, text="[copiar]",
                            font=("Segoe UI", 8), fg="#888", bg=BG, cursor="hand2")
        copy_lnk.pack(side="left", padx=(6,0))
        copy_lnk.bind("<Button-1>", lambda e: (inner.clipboard_clear(),
                                                inner.clipboard_append(account_url)))
        copy_lnk.bind("<MouseWheel>", _scroll)

        parrafo("2. En el panel derecho busca la seccion Clave de acceso de la API.")
        parrafo("3. Si no ves ninguna clave, hace click en Mostrar o Generar nueva clave.")
        parrafo("4. Copia esa clave y pegala en el campo API Key de Configuracion.")
        parrafo("   La API Key es personal e intransferible. No la compartas con nadie.")

        # ── CON O SIN ID_TICKET ──────────────────────────────
        seccion("Con o sin columna ID_Ticket")
        parrafo("La app ofrece dos formatos de Excel segun si usas un sistema de ticketera externo o no.")

        subtit("Excel CON columna ID_Ticket")
        parrafo("Ideal para equipos con ticketeras externas (ServiceNow, JIRA, SAP, etc.) donde los incidentes tienen un numero de ticket propio.")
        parrafo("El titulo del incidente en Redmine se conforma como: ID_Ticket - Titulo. Por ejemplo, si tu ticket externo es INC0012345 y el titulo es Descripcion del problema, el incidente en Redmine se llama: INC0012345 - Descripcion del problema. Esto permite trazabilidad entre el ticket del cliente y Redmine.")

        subtit("Excel SIN columna ID_Ticket")
        parrafo("Ideal para equipos sin ticketera externa. El titulo del incidente en Redmine se arma directamente con el Titulo que escribas, sin numero adicional adelante.")

        subtit("Cuando usar cada uno:")
        box("Con ID_Ticket",
            "Tenes numero de ticket del cliente (INC, REQ, CHG, etc.) de una ticketera externa.")
        box("Sin ID_Ticket",
            "No tenes ticketera externa o no necesitas rastrear por numero de ticket.")

        # ── ESCENARIOS ───────────────────────────────────────
        seccion("Escenarios de carga de horas")
        escenarios = [
            ("Escenario 1 — Tengo el ID de Redmine",
             "Completas ID_Redmine y Comentario. La app carga las horas directamente. Si ID_Ticket o Titulo estan vacios, los trae de Redmine y los completa en tu Excel."),
            ("Escenario 2 — No tengo ID de Redmine (issue nuevo)",
             "Completas Cliente, ID_Ticket (opcional), Titulo y Comentario. La app crea el incidente en Redmine, carga las horas y devuelve el ID_Redmine a tu Excel."),
            ("Escenario 3 — Sin Comentario",
             "Si dejas Comentario vacio, la app lo arma automaticamente como ID_Ticket - Titulo y lo escribe de vuelta en tu Excel."),
            ("Escenario 4 — Dos filas mismo incidente",
             "Si dos filas tienen el mismo Cliente + ID_Ticket + Titulo sin ID_Redmine, la app crea un solo incidente y carga ambas entradas de horas en el."),
        ]
        for t_esc, d_esc in escenarios:
            box(t_esc, d_esc)

        # ── CLIENTES ESPECIALES ──────────────────────────────
        seccion("Clientes especiales: Licencia y Vacaciones")
        parrafo("Estos clientes van siempre al incidente #7038, sin importar mayusculas ni minusculas:")
        box("Licencia",
            "Escribi licencia en la columna Cliente. Las horas van al incidente #7038.",
            color_bg="#fff8e1", color_titulo="#e67e00")
        box("Vacaciones",
            "Escribi vacaciones, vacacion o vacas en la columna Cliente. Las horas van al incidente #7038.",
            color_bg="#fff8e1", color_titulo="#e67e00")
        parrafo("No necesitas configurar nada en la pestana Clientes ni poner el ID_Redmine. La app lo resuelve automaticamente con el Comentario que hayas cargado.")




# ============================================================
#  ENTRY POINT
# ============================================================

if __name__ == "__main__":
    if not any("pyinstaller" in a.lower() for a in sys.argv):
        # Verificar instancia unica
        _lock_socket = verificar_instancia_unica()
        if _lock_socket is None:
            import tkinter as tk
            from tkinter import messagebox
            _root = tk.Tk()
            _root.withdraw()
            messagebox.showwarning(
                "Ya esta abierta",
                "El Cargador de Horas ya esta corriendo. Buscalo en la barra de tareas.")
            _root.destroy()
            sys.exit(0)

        auth = cargar_auth()
        if auth.get("hash"):
            LoginScreen(auth).mainloop()
        else:
            app = App()
            try:
                ico = get_asset("HM_Icono.ico")
                if os.path.exists(ico):
                    app.iconbitmap(ico)
                    app.after(200, lambda: app.iconbitmap(ico))
            except Exception:
                pass
            app.mainloop()
